#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Mazama_BP_2017.py

A script to convert the British Petroleum Statistical Review Excel workbook into
a series of standardized ASCII CSV files for ingest by other software.

"""

import sys
reload(sys)  # Reload does the trick!
sys.setdefaultencoding('UTF8')
from openpyxl import load_workbook
from Mazama_Countries import CountryTranslator
import pprint
import collections

RELEASE_YEAR = 2017
CSV_START_YEAR = 1965


########################################
# get_data_dictionary
# 
# Returns a data dictionary filled with the contents of a BP Stat Review data sheet.

def get_data_dictionary(sheet,rowrange,colrange):

    result = {}

    ct = CountryTranslator('BP_2015')
    # NOTE:  The following notes follow
    ignore_me = ["",
                 u"Oil: Production *",
                 u"* Includes crude oil, shale oil, oil sands and NGLs ( the liquid content of natural gas where this is recovered separately).",
                 u"Excludes liquid fuels from other sources such as biomass and coal derivatives.",
                 u"^ Less than 0.05.",
                 u"w Less than 0.05%.",
                 u"Excludes Former Soviet Union.",
                 u"# Excludes Estonia, Latvia and Lithuania prior to 1985 and Slovenia prior to 1991.",
                 u"Notes: Annual changes and shares of total are calculated using million tonnes per annum figures.",
                 u"* Includes crude oil, tight oil, oil sands and NGLs (the liquid content of natural gas where this is recovered separately). Excludes liquid fuels from other sources such as biomass and derivatives of coal and natural gas.",
                 u"coal and natural gas.",
                 u"n/a not available.",
                 u"Note: Annual changes and shares of total are calculated using million tonnes per annum figures.",
                 #
                 u"Oil: Consumption *",
                 u"* Includes crude oil, shale oil, oil sands and NGLs (the liquid content of natural gas where this is recovered separately).",
                 u"* Inland demand plus international aviation and marine bunkers and refinery fuel and loss.  Consumption of fuel ethanol and biodiesel is also included.",
                 u"* Includes crude oil, tight oil, oil sands and NGLs (the liquid content of natural gas where this is recovered separately). Excludes liquid fuels from other sources such as biomass and derivatives of",
                 u"* Inland demand plus international aviation and marine bunkers and refinery fuel and loss. Consumption of biogasoline (such as ethanol), biodiesel and derivatives of coal and natural gas are also included.",
                 u"Notes: Differences between these world consumption figures and world production statistics are accounted for by stock changes, consumption of non-petroleum additives",
                 u"* Excludes gas flared or recycled. Includes natural gas produced for Gas-to-Liquids transformation.",
                 u"Notes:  As far as possible, the data above represent standard cubic metres (measured at 15ºC and 1013 mbar); as they are derived directly from tonnes of oil equivalent using an average conversion factor,",
                 u"Notes: Annual changes and shares of total are calculated using million tonnes of oil equivalent figures.",
                 u"Notes: As the data above are derived from tonnes oil equivalent using average conversion factors, they do not necessarily equate with gas volumes expressed in specific national terms.",
                 u"Annual changes and shares of total are calculated using million tonnes of oil equivalent figures.",
                 u"*Excludes gas flared or recycled. Includes natural gas produced for Gas-to-Liquids transformation.",
                 u"Natural Gas: Consumption*",
                 u"* Excludes natural gas converted to liquid fuels but includes derivatives of coal as well as natural gas consumed in Gas-to-Liquids transformation.",
                 u"Notes: As far as possible, the data above represent standard cubic metres (measured at 15oC and 1013 mbar); as they are derived directly from tonnes",
                 u"Notes: The difference between these world consumption figures and the world production statistics is due to variations in stocks at storage facilities",
                 u"* Commercial solid fuels only, i.e. bituminous coal and anthracite (hard coal), and lignite and brown (sub-bituminous) coal, and other solid commercial solid fuels. Includes coal produced for Coal-to-Liquids and Coal-to-Gas transformations.",
                 u"^ Less than 0.5.",
                 u"Notes:  Annual changes and shares of total are calculated using million tonnes per annum figures.",
                 u"Differences between these world consumption figures and world production statistics are accounted for by stock changes, consumption of non-petroleum additives",
                 u"and substitute fuels, and unavoidable disparities in the definition, measurement or conversion of oil supply and demand data.",
                 u"Note: Differences between these world consumption figures and world production statistics are accounted for by stock changes, consumption of non-petroleum additives",
                 #
                 u"Natural Gas: Production *",
                 u"* Excluding gas flared or recyled.",
                 u"Notes:  As far as possible, the data above represent standard cubic metres measured at 15oC and 1013 millibar (mbar); as they are derived directly from tonnes",
u"Note: Annual changes and shares of total are calculated using million tonnes of oil equivalent figures.",
                 u"* Commercial solid fuels only, i.e. bituminous coal and anthracite (hard coal), and lignite and brown (sub-bituminous) coal and other commercial solid fuels. Includes coal produced for Coal-to-Liquids and Coal-to-Gas transformations.",
                 u"* Commercial solid fuels only, i.e. bituminous coal and anthracite (hard coal), and lignite and brown (sub-bituminous) coal. Excludes coal converted to liquid or gaseous fuels, but includes coal consumed in transformation processes.",
                 u"""Note:  Differences between these world consumption figures and the world production statistics are accounted for by stock changes, and unavoidable disparities in the definition, measurement
or conversion of coal supply and demand data.""",
                 u"Note: Annual changes and shares of total are calculated using million tonnes of oil equivalent figures.",
                 u"* Based on gross generation and not accounting for cross-border electricity supply. Converted on the basis of thermal equivalence assuming 38% conversion efficiency in a modern thermal power station.",
                 u"Note: Annual changes and shares of total are calculated using million tonnes of oil equivalent figures.",
                 u"Renewables: Consumption *",
                 u"* Based on gross generation from renewable sources including wind,geothermal,solar,biomass and waste, and not accounting for cross-border electricity supply.",
                 u"Renewables: Consumption *",
                 u"* Based on gross generation from renewable sources including wind,geothermal,solar,biomass and waste, and not accounting for cross-border electricity supply.",
                 u"* Based on gross generation and not accounting for cross-border electricity supply. Converted on the",
                 u"basis of thermal equivalence assuming 38% conversion efficiency in a modern thermal power station.",
                 u"of oil equivalent using an average conversion factor, they do not necessarily equate with gas volumes expressed in specific national terms.",
                 u"Notes: Annual changes and shares of total are calculated in million tonnes of oil equilevent figures.",
                 u"As the data above are derived from tonnes oil equivalent using average conversion factors,",
                 u"they do not necessarily equate with gas volumes expressed in specific national terms.",
                 u"w Less than 0.05%",
                 #
                 u"Natural Gas: Consumption",
                 u"Oil: Production*",
                 u"CIS",
                 u"* Includes crude oil, shale oil, oil sands and NGLs (natural gas liquids - the liquid content of natural gas where this is recovered separately).",
                 u"Excludes liquid fuels from other sources such as biomass and derivatives of coal and natural gas.",
                 u"# Excludes Estonia, Latvia and Lithuania prior to 1985 and Slovenia prior to 1990.",
                 u"Notes: Annual changes and shares of total are calculated using million tonnes of oil equilevent figures.",
                 u"As far as possible, the data above represent standard cubic metres (measured at 15oC and 1013 mbar); as they are derived directly from tonnes",
                 #
                 u"Natural Gas: Consumption",
                 u"Oil: Consumption*",
                 #u"Billion cubic feet",
                 u"Natural Gas: Production*",
                 u"Note: As far as possible, the data above represent standard cubic metres (measured at 15oC and 1013 mbar); as they are derived directly from tonnes",
                 u"The difference between these world consumption figures and the world production statistics is due to variations in stocks at storage facilities",
                 u"and liquefaction plants, together with unavoidable disparities in the definition, measurement or conversion of gas supply and demand data.",
                 u"Note: The difference between these world consumption figures and the world production statistics is due to variations in stocks at storage facilities",
                 #
                 u"Coal: Production *",
                 u"* Commercial solid fuels only, i.e. bituminous coal and anthracite (hard coal), and lignite and brown (sub-bituminous) coal.",
                 u"Notes: Annual changes and shares of total are based on data expressed in tonnes oil equivalent.",
                 #
                 u"Coal: Consumption *",
                 #
                 u"Nuclear: Consumption *",
                 u"* Based on gross generation and not accounting for cross-border electricity supply.",
                 u"Converted on the basis of thermal equivalence assuming 38% conversion efficiency in a modern thermal power station.",
                 #
                 u"Hydroelectricity: Consumption *",
                 u"* Based on gross primary hydroelectric generation and not accounting for cross-border electricity supply.",
                 u"* Based on gross primary hydroelectric generation and not accounting for cross-border electricity supply.  Converted on the basis of thermal equivalence assuming 38%",                 
                 u"conversion efficiency in a modern thermal power station.",
                 #
                 u"Other renewables: Consumption *",
                 u"* Based on gross generation from renewable sources including wind,geothermal,solar,biomass and waste, and not accounting for cross border electricity supply.",
                 u"Converted on the basis of thermal equivalence assuming 38% conversion efficiency in a modern thermal power station.",
                 u"^ Less than 0.05",
                 u"♦ Less than 0.05%", # TODO:  This string doesn't match                
                 u"# Excludes Slovenia prior to 1991",
                 #
                 u"Renewables: Consumption - Solar *",
                 u"Other North America",
                 u"♦ Less than 0.05%.", # TODO:  This string doesn't match
                 #
                 u"Renewables: Consumption - Wind *",
                 u"Coal: Production*",
                 u"* Commercial solid fuels only, i.e. bituminous coal and anthracite (hard coal), and lignite and brown (sub-bituminous) coal, and other commercial solid fuels. Includes coal produced for Coal-to-Liquids and Coal-to-Gas transformations.",
                 u"Renewables: Consumption  - Solar*",
                 u"Renewables: Consumption  - Wind*",
                 u"Nuclear: Consumption*",
                 u"Coal: Consumption*",
                 u"* Commercial solid fuels only, i.e. bituminous coal and anthracite (hard coal), and lignite and brown (sub-bituminous) coal, and other commercial solid fuels.",
                 u"Excludes coal converted to liquid or gaseous fuels, but includes coal consumed in transformation processes.",
                 u"Hydroelectricity: Consumption*",
                 u"* Based on gross primary hydroelectric generation and not accounting for cross-border electricity supply. Converted on the",
                 u"Renewables: Consumption*",
                  u"* Based on gross generation from renewable sources including wind, geothermal, solar, biomass and waste, and not accounting for cross-border electricity supply.",
                 u"Renewables: Consumption - Solar*",
                 u"Renewables: Consumption - Wind*",
                 u"Natural Gas: Consumption in billion cubic metres*",
                 ]

    for r in rowrange:
        BP_country_code = sheet.cell(row=r,column=1+0).value
        if BP_country_code == None:
            continue
        # NOTE:  Remove pound sterling sign (u'\xa3') from 'Non-OPEC'
        # NOTE:  Strip initial and trailing space
        BP_country_code = BP_country_code.replace(u'\xa3','').strip()
        if BP_country_code in ignore_me:
            continue
        try:
            MZM_code = ct.get_MZM_code(BP_country_code)
            ###print("DEBUG: MZM_code = '%s'" % (MZM_code))
            result[MZM_code] = []
            
            for c in colrange:
                value = sheet.cell(row=r,column=c).value
                data_type = sheet.cell(row=r,column=c).data_type
                
                if data_type == 'n':
                    if r == 2:
                        result["YEAR"].append(value)
                    else:
                        result[MZM_code].append(value)
                        
                elif data_type == 's':
                    if value == u'-':
                        result[MZM_code].append(0.0)
                    elif value == u'^':
                        result[MZM_code].append(0.0)
                    elif value == u'n/a':
                        result[MZM_code].append("na")
                    else:
                        try:
                            value = float(value)
                            result[MZM_code].append(value)
                        except Exception, e:
                            print("ERROR: " + str(e) + ": cell value \"" + value + "\" is not handled.")
                            sys.exit(1)
                    
                else:
                    print("UNKNOWN data_type %d" % (data_type))
                    sys.exit(1)
            
        except Exception, e:
            ###error_text = "ERROR in get_data_dictionary:  BP_country_code = \n\t%s\nerror = \n\t%s" % (BP_country_code,e)
            ###error_text = "ERROR: %s\nu\"%s\"," % (e, BP_country_code)
            error_text = "                 u\"%s\"," % (BP_country_code)
            print(error_text)
            ###exit(1)

            pass

    return(result)


########################################
# write_data_as_csv
#
# The csv file is organized as Year (row) X MZM_code (col)

def write_data_as_csv(filename, Data, rowrange, rounding, data_start_year=1965):

    # Get a sorted list of keys
    MZM_codes = Data.keys()
    MZM_codes.sort()
    MZM_codes.remove("YEAR")

    # Write out the header line
    filename.write("\"YEAR\"")
    for MZM_code in MZM_codes:
        filename.write(",\"" + MZM_code + "\"")
    filename.write("\n")

    # Prefill with 'na' as needed
    for row in range(CSV_START_YEAR, data_start_year):
        filename.write(str(row))
        for MZM_code in MZM_codes:
            filename.write(",\"na\"")
        filename.write("\n")

    # For each row, continue by writing out the year and all values
    for row in rowrange:
        filename.write(str(Data["YEAR"][row]))
        for MZM_code in MZM_codes:
            try:
                filename.write("," + str(round(Data[MZM_code][row],rounding)))
            except TypeError:
                filename.write(",\"na\"")
        filename.write("\n")


##############################################################################
# Main program
#
def main():

    stat_review = 'BP_2017.xlsx'

    print("Loading %s ..." % (stat_review))

    try:
        workbook = load_workbook(filename=stat_review)
    except:
        print("*** Open failed: %s: %s" % (sys.exc_info()[:2]))

    print("Successfully opened workbook.")

    title_to_filename = {
        #'Contents',
        #'Primary Energy Consumption',
        #'Primary Energy - Cons by fuel',
        #'Oil - Proved reserves',
        #'Oil - Proved reserves history',
        'Oil Production - Barrels' : "BP_2017_oil_production_bbl",
        'Oil Production - Tonnes' : "BP_2017_oil_production_mtoe",
        'Oil Consumption -  Barrels': "BP_2017_oil_consumption_bbl",
        'Oil Consumption - Tonnes': "BP_2017_oil_consumption_mtoe",
        #'Oil - Regional Consumption ',
        #'Oil - Spot crude prices',
        #'Oil - Crude prices since 1861',
        #'Oil - Refinery throughput',
        #'Oil - Refinery capacities',
        #'Oil - Regional refining margins',
        #'Oil - Trade movements',
        #'Oil - Inter-area movements ',
        #'Oil - Trade 2015 - 2016',
        #'Gas - Proved reserves',
        #'Gas - Proved reserves history ',
        'Gas Production - Bcm' : "BP_2017_gas_production_m3",
        'Gas Production - Bcf' : "BP_2017_gas_production_ft3",
        'Gas Production - Mtoe' : "BP_2017_gas_production_mtoe",
        'Gas Consumption - Bcm' : "BP_2017_gas_consumption_m3",
        'Gas Consumption - Bcf' : "BP_2017_gas_consumption_ft3",
        'Gas Consumption - Mtoe' : "BP_2017_gas_consumption_mtoe",
        # 'Gas - Trade - pipeline',
        # 'Gas - Trade movements LNG',
        # 'Gas - Trade 2015-2016',
        # 'Gas - Prices ',
        # 'Coal - Reserves',
        # 'Coal - Prices',
        'Coal Production - Tonnes' : "BP_2017_coal_production_ton",
        'Coal Production - Mtoe' : "BP_2017_coal_production_mtoe",
        'Coal Consumption -  Mtoe' : "BP_2017_coal_consumption_mtoe",
        'Nuclear Consumption - TWh' : "BP_2017_nuclear_consumption_twh",
        'Nuclear Consumption - Mtoe' : "BP_2017_nuclear_consumption_mtoe",
        'Hydro Consumption - TWh' : "BP_2017_hydro_consumption_twh",
        'Hydro Consumption - Mtoe': "BP_2017_hydro_consumption_mtoe",
        'Other renewables -TWh' : "BP_2017_renewables_consumption_twh",
        'Other renewables - Mtoe' : "BP_2017_renewables_consumption_mtoe",
        'Solar Consumption - TWh' : "BP_2017_solar_consumption_twh",
        'Solar Consumption - Mtoe': "BP_2017_solar_consumption_mtoe",
        'Wind Consumption - TWh ': "BP_2017_wind_consumption_twh",
        'Wind Consumption - Mtoe': "BP_2017_wind_consumption_mtoe",
        # 'Geo Biomass Other - TWh',
        # 'Geo Biomass Other - Mtoe',
        # 'Biofuels Production - Kboed',
        # 'Biofuels Production - Ktoe',
        # 'Electricity Generation ',
        # 'Carbon Dioxide Emissions',
        # 'Geothermal capacity',
        # 'Solar capacity',
        # 'Wind capacity',
        # 'Approximate conversion factors',
        # 'Definitions'
    }

    print "Verifying expected worksheets is present ..."
    title_to_worksheet = dict([(sheet.title, sheet) for sheet in workbook.worksheets])
    for title in title_to_filename.keys():
        if title in title_to_worksheet:
            print "%s => OK" % (title)
        else:
            pprint.pprint([sheet.title for sheet in workbook.worksheets])
            print "missing worksheet %s" % title
            sys.exit(1)

    import collections

    #sheet_indices = [sheet_indices[8]]
    for title in title_to_filename.keys():
        file_name = title_to_filename[title] + ".csv"
        sheet = title_to_worksheet[title]
        sheet_title = sheet.cell(row=1+0,column=1+0).value.replace('*','').rstrip()
        title = sheet.cell(row=1+0,column=1+0).value.replace('*','').rstrip()
        print "title => %s => %s" % (title, sheet_title)
        units = sheet.cell(row=1+2,column=1+0).value.lower()

        # Determine rows and columns to read
        data_start_year = sheet.cell(row=1+2,column=1+1).value
        col_hi = RELEASE_YEAR - data_start_year + 1
        colrange = range(1+1,col_hi+1)
        rowrange = range(1,100) # Rowrange is larger than needed and rows that don't have country names will be skipped

        print ("Converting %s (%s)" % (title,units)).ljust(85),
        #file_name = file_name.replace("_2017_", "_")
        #file_name = file_name.replace("_renewables_", "_other_renewables_")
        print "=> %s ..." % (file_name)
        file = open(file_name,'w')
        file.write("title         = ASCII CSV version of worksheet \"%s\" from the 2017 British Petroleum Statistical Review\n" % (title))
        file.write("file URL      = http://mazamascience.com/Data/Energy/BP/2017/%s\n" % (file_name))
        file.write("original data = http://www.bp.com/content/dam/bp/en/corporate/excel/energy-economics/statistical-review-2017/bp-statistical-review-of-world-energy-2017-underpinning-data.xlsx\n")
        file.write("country codes = ISO3166-1 two-letter codes or 'BP_~~~' for non-standard BP groupings (e.g. BP_TNA = Total North America)\n")
        file.write("units         = %s\n" % (units))
        file.write("\n")
        Data = get_data_dictionary(sheet,rowrange,colrange)
        rowrange = range(0,col_hi-1)
        rounding = 3
        write_data_as_csv(file,Data,rowrange,rounding,data_start_year)
        file.close()

################################################################################

if __name__ == "__main__":
    main()

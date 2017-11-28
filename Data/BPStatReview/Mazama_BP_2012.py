#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Mazama_BP_2012.py

A script to convert the British Petroleum Statistical Review Excel workbook into
a series of standardized ASCII CSV files for ingest by other software.

"""

import sys
from openpyxl import load_workbook
from Mazama_Countries import CountryTranslator

RELEASE_YEAR = 2012
CSV_START_YEAR = 1965


########################################
# get_data_dictionary
# 
# Returns a data dictionary filled with the contents of a BP Stat Review data sheet.

def get_data_dictionary(sheet,rowrange,colrange):

    result = {}

    ct = CountryTranslator('BP_2012')
      
    # NOTE:  The following notes follow  
    ignore_me = ["",
                 u"Oil: Production *",
                 u"* Includes crude oil, shale oil, oil sands and NGLs ( the liquid content of natural gas where this is recovered separately).",
                 u"Excludes liquid fuels from other sources such as biomass and coal derivatives.",
                 u"^ Less than 0.05.",
                 u"w Less than 0.05%.",
                 u"Excludes Former Soviet Union.",
                 u"# Excludes Estonia, Latvia and Lithuania prior to 1985 and Slovenia prior to 1991.",
                 u"Notes: Annual changes and shares of total are calculated using million tonnes per annum figures.",
                 #
                 u"Oil: Consumption *",
                 u"* Includes crude oil, shale oil, oil sands and NGLs (the liquid content of natural gas where this is recovered separately).",
                 u"* Inland demand plus international aviation and marine bunkers and refinery fuel and loss.  Consumption of fuel ethanol and biodiesel is also included.",
                 u"^ Less than 0.5.",
                 u"Notes:  Annual changes and shares of total are calculated using million tonnes per annum figures.",
                 u"Differences between these world consumption figures and world production statistics are accounted for by stock changes, consumption of non-petroleum additives",
                 u"and substitute fuels, and unavoidable disparities in the definition, measurement or conversion of oil supply and demand data.",
                 u"Note: Differences between these world consumption figures and world production statistics are accounted for by stock changes, consumption of non-petroleum additives",
                 #
                 u"Natural Gas: Production *",
                 u"* Excluding gas flared or recyled.",
                 u"Notes:  As far as possible, the data above represent standard cubic metres measured at 15oC and 1013 millibar (mbar); as they are derived directly from tonnes",
                 u"of oil equivalent using an average conversion factor, they do not necessarily equate with gas volumes expressed in specific national terms.",
                 u"Notes: Annual changes and shares of total are calculated in million tonnes of oil equilevent figures.",
                 u"As the data above are derived from tonnes oil equivalent using average conversion factors,",
                 u"they do not necessarily equate with gas volumes expressed in specific national terms.",
                 u"w Less than 0.05%",
                 #
                 u"Natural Gas: Consumption",
                 u"Notes: Annual changes and shares of total are calculated using million tonnes of oil equilevent figures.",
                 u"As far as possible, the data above represent standard cubic metres (measured at 15oC and 1013 mbar); as they are derived directly from tonnes",
                 #
                 u"Natural Gas: Consumption",
                 u"The difference between these world consumption figures and the world production statistics is due to variations in stocks at storage facilities",
                 u"and liquefaction plants, together with unavoidable disparities in the definition, measurement or conversion of gas supply and demand data.",
                 u"Note: The difference between these world consumption figures and the world production statistics is due to variations in stocks at storage facilities",
                 #
                 u"Coal: Production *",
                 u"* Commercial solid fuels only, i.e. bituminous coal and anthracite (hard coal), and lignite and brown (sub-bituminous) coal.",
                 u"Notes: Annual changes and shares of total are based on data expressed in tonnes oil equivalent.",
                 #
                 u"Coal: Consumption *",
                 #
                 u"Nuclear: Consumption *",
                 u"* Based on gross generation and not accounting for cross-border electricity supply.",
                 u"Converted on the basis of thermal equivalence assuming 38% conversion efficiency in a modern thermal power station.",
                 #
                 u"Hydroelectricity: Consumption *",
                 u"* Based on gross primary hydroelectric generation and not accounting for cross-border electricity supply.",
                 u"* Based on gross primary hydroelectric generation and not accounting for cross-border electricity supply.  Converted on the basis of thermal equivalence assuming 38%",                 
                 u"conversion efficiency in a modern thermal power station.",
                 #
                 u"Other renewables: Consumption *",
                 u"* Based on gross generation from renewable sources including wind,geothermal,solar,biomass and waste, and not accounting for cross border electricity supply.",
                 u"Converted on the basis of thermal equivalence assuming 38% conversion efficiency in a modern thermal power station.",
                 u"^ Less than 0.05",
                 u"♦ Less than 0.05%", # TODO:  This string doesn't match                
                 u"# Excludes Slovenia prior to 1991",
                 #
                 u"Renewables: Consumption - Solar *",
                 u"Other North America",
                 u"♦ Less than 0.05%.", # TODO:  This string doesn't match
                 #
                 u"Renewables: Consumption - Wind *",                 
                 ]

    for r in rowrange:
        BP_country_code = sheet.cell(row=r,column=0).value
        if BP_country_code == None:
            continue
        # NOTE:  Remove pound sterling sign (u'\xa3') from 'Non-OPEC'
        # NOTE:  Strip initial and trailing space
        BP_country_code = BP_country_code.replace(u'\xa3','').strip()
        if BP_country_code in ignore_me:
            continue
        try:
            MZM_code = ct.get_MZM_code(BP_country_code)
            ###print("DEBUG: MZM_code = '%s'" % (MZM_code))
            result[MZM_code] = []
            
            for c in colrange:
                value = sheet.cell(row=r,column=c).value
                data_type = sheet.cell(row=r,column=c).data_type
                
                if data_type == 'n':
                    if r == 2:
                        result["YEAR"].append(value)
                    else:
                        result[MZM_code].append(value)
                        
                elif data_type == 's':
                    if value == u'-':
                        result[MZM_code].append(0.0)
                    elif value == u'^':
                        result[MZM_code].append(0.0)
                    elif value == u'n/a':
                        result[MZM_code].append("na")
                    else:
                        try:
                            value = float(value)
                            result[MZM_code].append(value)
                        except Exception, e:
                            print("ERROR: " + str(e) + ": cell value \"" + value + "\" is not handled.")
                            sys.exit(1)
                    
                else:
                    print("UNKNOWN data_type %d" % (data_type))
                    sys.exit(1)
            
        except Exception, e:
            ###error_text = "ERROR in get_data_dictionary:  BP_country_code = \n\t%s\nerror = \n\t%s" % (BP_country_code,e)
            ###error_text = "ERROR: %s\nu\"%s\"," % (e, BP_country_code)
            error_text = "                 u\"%s\"," % (BP_country_code)
            print(error_text)
            ###exit(1)

            pass

    return(result)


########################################
# write_data_as_csv
#
# The csv file is organized as Year (row) X MZM_code (col)

def write_data_as_csv(filename, Data, rowrange, rounding, data_start_year=1965):

    # Get a sorted list of keys
    MZM_codes = Data.keys()
    MZM_codes.sort()
    MZM_codes.remove("YEAR")

    # Write out the header line
    filename.write("\"YEAR\"")
    for MZM_code in MZM_codes:
        filename.write(",\"" + MZM_code + "\"")
    filename.write("\n")

    # Prefill with 'na' as needed
    for row in range(CSV_START_YEAR, data_start_year):
        filename.write(str(row))
        for MZM_code in MZM_codes:
            filename.write(",\"na\"")
        filename.write("\n")

    # For each row, continue by writing out the year and all values
    for row in rowrange:
        filename.write(str(Data["YEAR"][row]))
        for MZM_code in MZM_codes:
            try:
                filename.write("," + str(round(Data[MZM_code][row],rounding)))
            except TypeError:
                filename.write(",\"na\"")
        filename.write("\n")


##############################################################################
# Main program
#
def main():

    # NOTE:  The original 2012 BP file was called 'statistical_review_of_world_energy_full_report_2012.xlsx'
    stat_review = 'BP_2012.xlsx'

    print("Loading %s ..." % (stat_review))

    try:
        workbook = load_workbook(filename=stat_review)
    except:
        print("*** Open failed: %s: %s" % (sys.exc_info()[:2]))

    print("Successfully opened workbook.")

    # A list of file names, one for each worksheet
    csv_file_names = ["Contents", #0
                      "Oil - proved reserves",
                      "Oil - proved reserves history",
                      "BP_2012_oil_production_bbl",
                      "BP_2012_oil_production_mtoe",
                      "BP_2012_oil_consumption_bbl",
                      "BP_2012_oil_consumption_mtoe",
                      "Oil - Regional consumption ",
                      "Oil –  Spot crude prices",
                      "Oil - crude prices since 1861",
                      "Oil - Refinery capacities", # 10
                      "Oil - Refinery throughputs",
                      "Oil - Regional refining margins",
                      "Oil - Trade movements",
                      "Oil - Inter-area movements ",
                      "Oil - Imports and exports",
                      "Gas – Proved reserves",
                      "Gas - Proved reserves history ",
                      "BP_2012_gas_production_m3",
                      "BP_2012_gas_production_ft3",
                      "BP_2012_gas_production_mtoe", # 20
                      "BP_2012_gas_consumption_m3",
                      "BP_2012_gas_consumption_ft3",
                      "BP_2012_gas_consumption_mtoe",
                      "Gas - Trade - pipeline",
                      "Gas – Trade movements LNG",
                      "Gas - Trade 2010-2011",
                      "Gas - Prices ",
                      "Coal - Reserves",
                      "Coal - Prices",
                      "BP_2012_coal_production_ton", # 30
                      "BP_2012_coal_production_mtoe",
                      "BP_2012_coal_consumption_mtoe",
                      "BP_2012_nuclear_consumption_twh",
                      "BP_2012_nuclear_consumption_mtoe",
                      "BP_2012_hydro_consumption_twh",
                      "BP_2012_hydro_consumption_mtoe",
                      "BP_2012_renewables_consumption_mtoe",
                      "BP_2012_renewables_consumption_twh",
                      "BP_2012_solar_consumption_twh",
                      "BP_2012_wind_consumption_twh ", # 40
                      "Other consumption - Twh",
                      "Biofuels Production - barrels ",
                      "Biofuels Production - Ktoe",
                      "Primary Energy - Consumption",
                      "Primary Energy - Cons by fuel",
                      "Electricity Generation ",
                      "Carbon Dioxide Emissions",
                      "Geothermal capacity",
                      "Solar capacity",
                      "Wind capacity", # 50
                      "Approximate conversion factors",
                      "Definitions                     ",
                      ]
    # TODO:  A python OrderdDict would be better here once we start using python 2.7.
    # Work on all worksheets that are organized Country (row) by Year (col)
    # Note:  python indices are one less than Excel sheet indices
    sheet_indices = [3,4,5,6, # Oil Production & Consumption
                    18,19,20,21,22,23, # Gas Production & Consumption
                    30,31,32, # Coal Production & Consumption
                    33,34, # Nuclear Consumption
                    35,36, # Nuclear Consumption
                    37,38, # Other renewables
                    39,40, # Solar & Wind
                    ]
    
    for sheet_index in sheet_indices:
        
        sheet = workbook.worksheets[sheet_index]
        
        # Access and clean up title and units
        title = sheet.cell(row=0,column=0).value.replace('*','').rstrip()
        units = sheet.cell(row=2,column=0).value.lower()
        
        # Determine rows and columns to read
        data_start_year = sheet.cell(row=2,column=1).value
        col_hi = RELEASE_YEAR - data_start_year + 1
        colrange = range(1,col_hi)    
        rowrange = range(0,100) # Rowrange is larger than needed and rows that don't have country names will be skipped
        
        print("Converting %s (%s) Worksheet" % (title,units))
        
        file_name = csv_file_names[sheet_index] + ".csv"
        file = open(file_name,'w')
        file.write("title         = ASCII CSV version of worksheet %d, \"%s\" from the 2012 British Petroleum Statistical Review\n" % (sheet_index+1, title))
        file.write("file URL      = http://mazamascience.com/Data/Energy/BP/2012/%s\n" % (file_name))
        file.write("original data = http://www.bp.com/assets/bp_internet/globalbp/globalbp_uk_english/reports_and_publications/statistical_energy_review_2012/STAGING/local_assets/spreadsheets/statistical_review_of_world_energy_full_report_2012.xlsx\n")
        file.write("country codes = ISO3166-1 two-letter codes or 'BP_~~~' for non-standard BP groupings (e.g. BP_TNA = Total North America)\n")
        file.write("units         = %s\n" % (units))
        file.write("\n")
      
        Data = get_data_dictionary(sheet,rowrange,colrange)
        
        rowrange = range(0,col_hi-1)
        rounding = 3
        write_data_as_csv(file,Data,rowrange,rounding,data_start_year)
      
        file.close()       
        

################################################################################

if __name__ == "__main__":
  main()
